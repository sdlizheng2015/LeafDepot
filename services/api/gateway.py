"""
Gateway服务 - 向后兼容入口
"""
import warnings
import asyncio
import os
import time
import json
import logging
import glob
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime, timedelta

from fastapi import FastAPI, Request, BackgroundTasks, HTTPException, status, Body
from fastapi.responses import JSONResponse, Response
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import requests
import uvicorn
import sys

import xlsxwriter as XLSX  # 如果还没有导入的话

# 配置日志
# 创建 debug 日志目录（在文件开头就定义，供后续使用）
_project_root = Path(__file__).parent.parent.parent
_debug_log_dir = _project_root / "debug"
_debug_log_dir.mkdir(parents=True, exist_ok=True)

# 创建日志文件路径（按日期命名）
_log_filename = _debug_log_dir / \
    f"gateway_{datetime.now().strftime('%Y%m%d')}.log"

# 配置根日志记录器
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
    handlers=[
        logging.StreamHandler(),  # 控制台输出
        logging.FileHandler(str(_log_filename), encoding='utf-8')  # 文件输出
    ]
)
logger = logging.getLogger(__name__)
logger.info(f"日志文件保存路径: {_log_filename}")

# 导出供其他函数使用
debug_log_dir = _debug_log_dir

# 操作记录存储根目录
OPERATION_LOGS_DIR = debug_log_dir / "operation_logs"
OPERATION_LOGS_DIR.mkdir(parents=True, exist_ok=True)

# 创建各类型子目录
LOG_TYPES = ["inventory", "user_login",
             "user_management", "system_cleanup", "other"]
for log_type in LOG_TYPES:
    (OPERATION_LOGS_DIR / log_type).mkdir(parents=True, exist_ok=True)


class OperationLog(BaseModel):
    """操作记录模型"""
    id: str  # 格式: YYYYMMDD_HHMMSS_xxx
    timestamp: str  # ISO格式时间戳
    operation_type: str  # 操作类型
    user_id: Optional[str] = None  # 操作用户ID
    user_name: Optional[str] = None  # 操作用户名
    action: str  # 具体动作
    target: Optional[str] = None  # 操作目标（如任务ID、用户名）
    status: str  # 成功/失败/进行中
    details: Dict[str, Any] = {}  # 详细数据
    ip_address: Optional[str] = None  # 操作IP
    metadata: Dict[str, Any] = {}  # 元数据


def generate_operation_id() -> str:
    """生成操作记录ID"""
    return datetime.now().strftime("%Y%m%d_%H%M%S_") + str(int(time.time() * 1000))[-3:]


def save_operation_log(operation_log: OperationLog):
    """保存操作记录到文件"""
    try:
        # 根据类型选择目录
        log_type = operation_log.operation_type
        if log_type not in LOG_TYPES:
            log_type = "other"

        log_dir = OPERATION_LOGS_DIR / log_type

        # 创建文件名
        filename = f"{operation_log.id}.json"
        filepath = log_dir / filename

        # 保存为JSON
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(operation_log.dict(), f, ensure_ascii=False, indent=2)

        logger.info(f"操作记录已保存: {filepath}")
        return True
    except Exception as e:
        logger.error(f"保存操作记录失败: {str(e)}")
        return False


def log_operation(
    operation_type: str,
    action: str,
    user_id: Optional[str] = None,
    user_name: Optional[str] = None,
    target: Optional[str] = None,
    status: str = "success",
    details: Optional[Dict[str, Any]] = None,
    ip_address: Optional[str] = None,
    metadata: Optional[Dict[str, Any]] = None
):
    """记录操作的辅助函数"""
    operation_log = OperationLog(
        id=generate_operation_id(),
        timestamp=datetime.now().isoformat(),
        operation_type=operation_type,
        user_id=user_id,
        user_name=user_name,
        action=action,
        target=target,
        status=status,
        details=details or {},
        ip_address=ip_address,
        metadata=metadata or {}
    )

    return save_operation_log(operation_log)


def get_recent_operations(limit: int = 5, days: int = 180) -> List[Dict[str, Any]]:
    """
    获取最近的操作记录

    Args:
        limit: 返回数量限制
        days: 查询天数（默认180天，约6个月）
    """
    try:
        all_operations = []
        cutoff_date = datetime.now() - timedelta(days=days)

        # 遍历所有类型目录
        for log_type in LOG_TYPES:
            log_dir = OPERATION_LOGS_DIR / log_type
            if not log_dir.exists():
                continue

            # 查找所有JSON文件
            json_files = list(log_dir.glob("*.json"))

            for json_file in json_files:
                try:
                    with open(json_file, 'r', encoding='utf-8') as f:
                        data = json.load(f)

                    # 检查时间是否在有效期内
                    log_time_str = data.get("timestamp", "")
                    if log_time_str:
                        try:
                            log_time = datetime.fromisoformat(
                                log_time_str.replace('Z', '+00:00'))
                            if log_time < cutoff_date:
                                continue  # 跳过过期记录
                        except:
                            pass  # 如果时间解析失败，仍然包含

                    all_operations.append(data)
                except Exception as e:
                    logger.warning(f"读取操作记录文件失败 {json_file}: {str(e)}")
                    continue

        # 按时间倒序排序
        all_operations.sort(key=lambda x: x.get("timestamp", ""), reverse=True)

        # 限制返回数量
        return all_operations[:limit]

    except Exception as e:
        logger.error(f"获取操作记录失败: {str(e)}")
        return []


def get_all_operations(days: int = 180) -> List[Dict[str, Any]]:
    """获取指定天数内的所有操作记录"""
    try:
        all_operations = []
        cutoff_date = datetime.now() - timedelta(days=days)

        # 遍历所有类型目录
        for log_type in LOG_TYPES:
            log_dir = OPERATION_LOGS_DIR / log_type
            if not log_dir.exists():
                continue

            # 查找所有JSON文件
            json_files = list(log_dir.glob("*.json"))

            for json_file in json_files:
                try:
                    with open(json_file, 'r', encoding='utf-8') as f:
                        data = json.load(f)

                    # 检查时间是否在有效期内
                    log_time_str = data.get("timestamp", "")
                    if log_time_str:
                        try:
                            log_time = datetime.fromisoformat(
                                log_time_str.replace('Z', '+00:00'))
                            if log_time < cutoff_date:
                                continue  # 跳过过期记录
                        except:
                            pass  # 如果时间解析失败，仍然包含

                    all_operations.append(data)
                except Exception as e:
                    logger.warning(f"读取操作记录文件失败 {json_file}: {str(e)}")
                    continue

        # 按时间倒序排序
        all_operations.sort(key=lambda x: x.get("timestamp", ""), reverse=True)
        return all_operations

    except Exception as e:
        logger.error(f"获取所有操作记录失败: {str(e)}")
        return []


# 数据模型定义
class TaskStatus(BaseModel):
    """任务状态模型"""
    task_no: str
    status: str  # init, running, completed, failed
    current_step: int = 0
    total_steps: int = 0
    start_time: Optional[str] = None
    end_time: Optional[str] = None


class BinLocationStatus(BaseModel):
    """储位状态模型"""
    bin_location: str
    status: str  # pending, running, completed, failed
    sequence: int = 0
    image_data: Optional[Dict[str, Any]] = None
    compute_result: Optional[Dict[str, Any]] = None
    capture_time: Optional[str] = None
    compute_time: Optional[str] = None
    detect_result: Optional[Dict[str, Any]] = None  # Detect模块识别结果
    barcode_result: Optional[Dict[str, Any]] = None  # Barcode模块识别结果
    recognition_time: Optional[str] = None  # 识别时间


class InventoryTaskProgress(BaseModel):
    """盘点任务进度模型"""
    task_no: str
    status: str
    current_step: int
    total_steps: int
    progress_percentage: float
    bin_locations: List[BinLocationStatus]
    start_time: Optional[str] = None
    end_time: Optional[str] = None


# 请求模型
class ScanAndRecognizeRequest(BaseModel):
    """扫码+识别请求模型"""
    taskNo: str  # 任务编号
    binLocation: str  # 库位号
    pile_id: int = 1  # 堆垛ID，默认为1
    code_type: str = "ucc128"  # 条码类型，默认ucc128


class FrontendLogRequest(BaseModel):
    """前端日志请求模型"""
    level: str  # log, info, warn, error
    message: str
    timestamp: Optional[str] = None
    source: Optional[str] = None  # 前端来源标识
    extra: Optional[Dict[str, Any]] = None  # 额外信息


# 创建 FastAPI 应用实例
app = FastAPI(
    title="LeafDepot API Gateway",
    description="LeafDepot 系统 API 网关服务",
    version="1.0.0"
)

# 配置 CORS
# 从环境变量读取允许的源，如果没有设置则使用默认列表
cors_origins_env = os.getenv("CORS_ORIGINS", "")
if cors_origins_env:
    # 如果设置了环境变量，使用环境变量的值（逗号分隔）
    origins = [origin.strip() for origin in cors_origins_env.split(",")]
else:
    # 默认允许的源列表
    origins = [
        "http://localhost",
        "http://localhost:8000",
        "http://localhost:8001",
        "http://localhost:8080",  # 测试页面服务器端口
        "http://localhost:3000",
        "http://localhost:5000",
        "http://127.0.0.1:8000",
        "http://127.0.0.1:8001",
        "http://127.0.0.1:8080",  # 测试页面服务器端口
        "http://127.0.0.1:3000",
        "http://127.0.0.1:5000",
    ]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"],  # 暴露所有头部
    max_age=3600,  # 预检请求缓存时间
)

# Barcode功能开关（默认关闭）
ENABLE_BARCODE = os.getenv(
    "ENABLE_BARCODE", "false").lower() in ("true", "1", "yes")

# 导入条形码路由（如果存在且开关开启）
if ENABLE_BARCODE:
    try:
        from services.api.routers.barcode import router as barcode_router
        app.include_router(barcode_router)
        logger.info("条形码路由已启用")
    except ImportError:
        logger.warning("条形码路由模块未找到，跳过注册")
else:
    logger.info("条形码功能已禁用（ENABLE_BARCODE=false）")

# 导入工具模块

# 导入检测模块
try:
    from core.detection import count_boxes
    DETECT_MODULE_AVAILABLE = True
except ImportError as e:
    logger.warning(f"检测模块导入失败: {e}")
    DETECT_MODULE_AVAILABLE = False

# 导入条码识别模块（仅在开关开启时导入）
BARCODE_MODULE_AVAILABLE = False
if ENABLE_BARCODE:
    try:
        from core.vision.barcode_recognizer import BarcodeRecognizer
        BARCODE_MODULE_AVAILABLE = True
        logger.info("条形码识别模块已启用")
    except ImportError as e:
        logger.warning(f"条形码识别模块导入失败: {e}")
        BARCODE_MODULE_AVAILABLE = False
else:
    logger.info("条形码识别模块已禁用（ENABLE_BARCODE=false）")

# 配置常量（这些应该从配置文件或环境变量中读取）
# 默认使用模拟服务端口（6000），如果环境变量设置了则使用环境变量的值
LMS_BASE_URL = os.getenv("LMS_BASE_URL", "http://localhost:6000")
RCS_BASE_URL = os.getenv("RCS_BASE_URL", "http://localhost:8003")
RCS_PREFIX = os.getenv("RCS_PREFIX", "")

# 抓图脚本路径（应该从配置文件读取）
CAPTURE_SCRIPTS = [
    os.path.join(os.path.dirname(__file__), "..", "..",
                 "hardware", "cam_sys", "scan_1_capture.py"),
    os.path.join(os.path.dirname(__file__), "..", "..",
                 "hardware", "cam_sys", "scan_2_capture.py"),
]

# 延迟导入服务（避免循环导入）


def get_inventory_service():
    """获取盘点服务实例"""
    from services.vision.box_count_service import get_box_count_service
    return get_box_count_service()


# 任务状态存储（如果任务状态管理器不存在，使用简单的内存存储）
_inventory_tasks: Dict[str, TaskStatus] = {}
_inventory_task_bins: Dict[str, List[BinLocationStatus]] = {}
_inventory_task_details: Dict[str, Dict[str, Dict[str, Any]]] = {}

# 获取任务状态存储的辅助函数


def get_task_state_storage():
    """获取任务状态存储（优先使用任务状态管理器，否则使用简单内存存储）"""
    try:
        from services.api.state.task_state import get_task_state_manager
        manager = get_task_state_manager()
        return {
            "tasks": manager._inventory_tasks,
            "bins": manager._inventory_task_bins,
            "details": manager._inventory_task_details
        }
    except ImportError:
        # 如果任务状态管理器不存在，使用简单的内存存储
        return {
            "tasks": _inventory_tasks,
            "bins": _inventory_task_bins,
            "details": _inventory_task_details
        }


# 初始化模块级别的变量（供函数内部直接使用）
# 这些变量指向实际存储（通过 get_task_state_storage() 获取）
_storage = get_task_state_storage()
inventory_tasks = _storage["tasks"]
inventory_task_bins = _storage["bins"]
inventory_task_details = _storage["details"]

# 为了向后兼容，保留一些全局变量引用（通过延迟导入避免循环）


def __getattr__(name):
    """延迟导入以支持向后兼容"""
    if name == "inventory_tasks":
        try:
            from services.api.state.task_state import get_task_state_manager
            return get_task_state_manager()._inventory_tasks
        except ImportError:
            logger.warning("任务状态管理器模块不存在，使用简单内存存储")
            return _inventory_tasks
    elif name == "inventory_task_bins":
        try:
            from services.api.state.task_state import get_task_state_manager
            return get_task_state_manager()._inventory_task_bins
        except ImportError:
            logger.warning("任务状态管理器模块不存在，使用简单内存存储")
            return _inventory_task_bins
    elif name == "inventory_task_details":
        try:
            from services.api.state.task_state import get_task_state_manager
            return get_task_state_manager()._inventory_task_details
        except ImportError:
            logger.warning("任务状态管理器模块不存在，使用简单内存存储")
            return _inventory_task_details
    elif name in ["STATUS_KEY", "status_event", "robot_status_store"]:
        try:
            from services.api.state.robot_state import get_robot_state_manager
            manager = get_robot_state_manager()
            if name == "STATUS_KEY":
                from services.api.config import ROBOT_STATUS_KEY
                return ROBOT_STATUS_KEY
            elif name == "status_event":
                return manager._status_event
            elif name == "robot_status_store":
                return manager._robot_status_store
        except ImportError:
            logger.warning(f"机器人状态管理器模块不存在，无法访问 {name}")
            return None
    raise AttributeError(f"module '{__name__}' has no attribute '{name}'")


# ========== 在关键操作处添加记录 ==========
# 辅助函数：根据token获取用户信息
async def get_user_info_from_token(auth_token: str) -> Dict[str, Any]:
    """根据authToken获取用户信息"""
    try:
        # 调用LMS的auth/token接口
        lms_auth_url = f"{LMS_BASE_URL}/auth/token?token={auth_token}"
        response = requests.get(lms_auth_url)

        if response.status_code == 200:
            return response.json().get("data", {})
        else:
            return {}
    except Exception:
        return {}


# 1. 在登录接口添加记录
@app.post("/login")
async def login(request: Request):
    """处理前端登录请求，调用LMS的login接口"""
    try:
        # 从前端获取用户名和密码
        data = await request.json()
        username = data.get("username")
        password = data.get("password")

        # 验证输入
        if not username or not password:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="用户名和密码不能为空"
            )

        # 调用LMS的login接口
        lms_login_url = f"{LMS_BASE_URL}/login"
        headers = {
            "userCode": username,
            "password": password
        }
        logger.info(f"尝试连接LMS服务: {lms_login_url}")
        try:
            response = requests.get(lms_login_url, headers=headers, timeout=5)
        except requests.exceptions.ConnectionError as e:
            logger.error(f"无法连接到LMS服务 {lms_login_url}: {str(e)}")
            raise HTTPException(
                status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
                detail=f"无法连接到LMS服务，请确保LMS服务正在运行（{LMS_BASE_URL}）"
            )
        except requests.exceptions.Timeout:
            logger.error(f"连接LMS服务超时: {lms_login_url}")
            raise HTTPException(
                status_code=status.HTTP_504_GATEWAY_TIMEOUT,
                detail="LMS服务响应超时"
            )

        if response.status_code == 200:
            # 获取LMS返回的token
            lms_response = response.json()
            token = lms_response.get("authToken")

            if not token:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="登录成功但未返回authToken"
                )

            # 记录登录操作
            client_host = request.client.host if request.client else "unknown"
            log_operation(
                operation_type="user_login",
                action="用户登录",
                user_id=lms_response.get("userId"),
                user_name=lms_response.get("userName"),
                status="success",
                ip_address=client_host,
                details={
                    "login_method": "password",
                    "user_level": lms_response.get("userLevel")
                }
            )

            # 返回给前端的响应
            return {
                "success": True,
                "data": {
                    "userId": lms_response.get("userId"),
                    "userCode": lms_response.get("userCode"),
                    "userName": lms_response.get("userName"),
                    "authToken": token,
                    "userLevel": lms_response.get("userLevel"),
                }
            }
        else:
            # 记录登录失败
            client_host = request.client.host if request.client else "unknown"
            log_operation(
                operation_type="user_login",
                action="用户登录",
                user_id=username,
                status="failed",
                ip_address=client_host,
                details={
                    "error": response.text[:200],
                    "status_code": response.status_code
                }
            )

            raise HTTPException(
                status_code=response.status_code,
                detail=f"LMS登录失败: {response.text}"
            )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"登录请求失败: {str(e)}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"登录请求处理失败: {str(e)}"
        )


async def submit_inventory_task(task_no: str, bin_locations: List[str]):
    """下发盘点任务，接收任务编号和储位名称列表"""
    try:

        logger.info(f"下发盘点任务: {task_no}, 储位: {bin_locations}")

        url = f"{RCS_BASE_URL}{RCS_PREFIX}/api/robot/controller/task/submit"
        headers = {
            "X-lr-request-id": "ldui",
            "Content-Type": "application/json"
        }

        # 构建targetRoute数组
        target_route = []
        for index, location in enumerate(bin_locations):
            route_item = {
                "seq": index,
                "type": "ZONE",
                "code": location,  # 使用储位名称作为目标区域
            }
            target_route.append(route_item)

        # 构建请求体 - 单个任务对象
        request_body = {
            "taskType": "PF-CTU-COMMON-TEST",
            "targetRoute": target_route
        }

        response = requests.post(
            url, json=request_body, headers=headers, timeout=30)

        if response.status_code == 200:
            response_data = response.json()

            if response_data.get("code") == "SUCCESS":
                logger.info(f"储位 {bin_locations} 已发送到机器人系统")
                return {"success": True, "message": "盘点任务已下发"}
        else:
            return {"success": False, "message": "盘点任务下发失败"}

    except Exception as e:
        logger.error(f"下发盘点任务失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"下发盘点任务失败: {str(e)}"
        )


async def continue_inventory_task():
    """继续盘点任务"""
    try:
        logger.info(f"继续执行盘点任务")

        url = f"{RCS_BASE_URL}{RCS_PREFIX}/api/robot/controller/task/extend/continue"
        headers = {
            "X-lr-request-id": "ldui",
            "Content-Type": "application/json"
        }

        # 构建请求体
        request_body = {
            "triggerType": "TASK",
            "triggerCode": "001"
        }

        response = requests.post(
            url, json=request_body, headers=headers, timeout=30)

        if response.status_code == 200:
            response_data = response.json()

            if response_data.get("code") == "SUCCESS":
                logger.info(f"继续执行盘点任务命令已发送到机器人系统")
                return {"success": True, "message": "盘点任务已继续"}
        else:
            return {"success": False, "message": "盘点任务下发失败"}

    except Exception as e:
        logger.error(f"继续盘点任务失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"继续盘点任务失败: {str(e)}"
        )


async def update_robot_status(method: str, data: Optional[Dict] = None):
    """更新机器人状态并触发事件"""
    try:
        from services.api.state.robot_state import get_robot_state_manager
        manager = get_robot_state_manager()
        from services.api.config import ROBOT_STATUS_KEY

        # 保存状态信息
        manager._robot_status_store[ROBOT_STATUS_KEY] = {
            "method": method,
            "timestamp": time.time(),
            "data": data or {}
        }

        logger.info(f"更新机器人状态: {method}")

        # 设置事件，通知等待的进程
        manager._status_event.set()
    except ImportError:
        logger.warning("机器人状态管理器模块不存在，无法更新状态")


async def wait_for_robot_status(expected_method: str, timeout: int = 300):
    """
    等待特定机器人状态的同步函数

    这个函数会阻塞直到收到期望的状态或超时
    """
    try:
        from services.api.state.robot_state import get_robot_state_manager
        from services.api.config import ROBOT_STATUS_KEY
        manager = get_robot_state_manager()

        logger.info(f"开始等待机器人状态: {expected_method}, 超时: {timeout}秒")

        start_time = time.time()

        # 清除事件，确保我们等待的是新的事件
        manager._status_event.clear()

        # 检查是否已经有期望的状态
        if ROBOT_STATUS_KEY in manager._robot_status_store:
            current_status = manager._robot_status_store[ROBOT_STATUS_KEY]
            if current_status.get("method") == expected_method:
                logger.info(f"已存在期望状态: {expected_method}")
                return current_status

        while True:
            try:
                # 等待事件被设置
                await asyncio.wait_for(manager._status_event.wait(), timeout=1.0)

                # 检查状态
                if ROBOT_STATUS_KEY in manager._robot_status_store:
                    current_status = manager._robot_status_store[ROBOT_STATUS_KEY]
                    logger.info(f"收到机器人状态: {current_status.get('method')}")

                    if current_status.get("method") == expected_method:
                        logger.info(f"收到期望状态: {expected_method}")
                        return current_status

                # 重置事件，准备下一次等待
                manager._status_event.clear()

            except asyncio.TimeoutError:
                # 检查是否总时间超时
                elapsed_time = time.time() - start_time
                if elapsed_time >= timeout:
                    logger.error(f"等待机器人状态超时: {expected_method}")
                    raise asyncio.TimeoutError(f"等待 {expected_method} 状态超时")

                # 继续等待
                continue
    except ImportError:
        logger.error("机器人状态管理器模块不存在，无法等待状态")
        raise


async def execute_capture_script(script_path: str, task_no: str, bin_location: str) -> Dict[str, Any]:
    """
    在指定 Conda 环境中执行单个抓图脚本

    Args:
        script_path: 脚本路径
        task_no: 任务编号
        bin_location: 储位名称
        conda_env: Conda 环境名称，默认为 'your_env_name'

    Returns:
        脚本执行结果
    """
    conda_env = "tobacco_env"
    try:
        logger.info(f"在 Conda 环境 '{conda_env}' 中执行抓图脚本: {script_path}")

        # 方法1: 使用 conda run 命令
        # 构建命令行参数
        cmd = ["conda", "run", "-n", conda_env, "python", script_path,
               "--task-no", task_no, "--bin-location", bin_location]

        # 方法2: 直接使用 conda 环境中的 python 路径（如果知道路径）
        # 假设你的 conda 环境路径是已知的
        # conda_python_path = f"/home/user/anaconda3/envs/{conda_env}/bin/python"
        # cmd = [conda_python_path, script_path, "--task-no", task_no, "--bin-location", bin_location]

        # 执行脚本
        process = await asyncio.create_subprocess_exec(
            *cmd,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE
        )

        # 等待脚本完成
        stdout, stderr = await process.communicate()

        # 解析结果
        result = {
            "script": os.path.basename(script_path),
            "conda_env": conda_env,
            "returncode": process.returncode,
            "stdout": stdout.decode('utf-8') if stdout else "",
            "stderr": stderr.decode('utf-8') if stderr else "",
            "success": process.returncode == 0
        }

        if process.returncode == 0:
            logger.info(f"脚本执行成功: {script_path} (环境: {conda_env})")
        else:
            logger.error(
                f"脚本执行失败: {script_path}, 错误: {stderr.decode('utf-8')}")

        return result

    except FileNotFoundError as e:
        logger.error(f"conda 命令未找到或 Conda 环境 '{conda_env}' 不存在: {str(e)}")
        return {
            "script": os.path.basename(script_path),
            "conda_env": conda_env,
            "returncode": -1,
            "stdout": "",
            "stderr": f"Conda 环境 '{conda_env}' 未找到或 conda 命令不可用",
            "success": False
        }
    except Exception as e:
        logger.error(f"执行脚本失败 {script_path}: {str(e)}")
        return {
            "script": os.path.basename(script_path),
            "conda_env": conda_env,
            "returncode": -1,
            "stdout": "",
            "stderr": str(e),
            "success": False
        }


async def capture_images_with_scripts(task_no: str, bin_location: str) -> List[Dict[str, Any]]:
    """
    按顺序执行三个抓图脚本

    Args:
        task_no: 任务编号
        bin_location: 储位名称

    Returns:
        所有脚本的执行结果
    """
    results = []

    for i, script_path in enumerate(CAPTURE_SCRIPTS, 1):
        logger.info(f"开始执行第 {i} 个抓图脚本: {script_path}")

        try:
            # 检查脚本文件是否存在
            if not os.path.exists(script_path):
                logger.error(f"脚本文件不存在: {script_path}")
                results.append({
                    "script": script_path,
                    "success": False,
                    "error": "脚本文件不存在"
                })
                continue

            # 执行脚本
            result = await execute_capture_script(script_path, task_no, bin_location)
            results.append(result)

            # 如果脚本执行失败，可以选择是否继续执行后续脚本
            if not result["success"]:
                logger.warning(f"第 {i} 个抓图脚本执行失败，继续执行下一个脚本")
                # 可以根据业务需求决定是否中断
                # continue

            # 脚本之间的短暂延迟（可选）
            if i < len(CAPTURE_SCRIPTS):
                await asyncio.sleep(0.5)

        except Exception as e:
            logger.error(f"执行第 {i} 个抓图脚本时发生异常: {str(e)}")
            results.append({
                "script": script_path,
                "success": False,
                "error": str(e)
            })

    return results


async def process_single_bin_location(task_no: str, bin_location: str, index: int, total: int):
    """处理单个储位的完整流程"""
    result = {
        "binLocation": bin_location,
        "sequence": index + 1,
        "startTime": datetime.now().isoformat(),
        "endTime": None,
        "status": None
    }

    try:
        # 更新任务状态（已在execute_inventory_workflow中处理）

        # 等待机器人就位
        logger.info(f"============等待机器人就位信息: {bin_location}")
        try:
            ctu_status = await wait_for_robot_status("end", timeout=300)

            # 这个判断一定会执行，因为wait_for_robot_status会阻塞直到收到end状态或超时
            if ctu_status and ctu_status.get("method") == "end":

                # 执行抓图脚本
                capture_results = await capture_images_with_scripts(task_no, bin_location)
                result["captureResults"] = capture_results

                # 检查抓图结果
                successful_scripts = sum(
                    1 for r in capture_results if r.get("success"))
                if successful_scripts < len(CAPTURE_SCRIPTS):
                    logger.warning(
                        f"部分抓图脚本执行失败: {successful_scripts}/{len(CAPTURE_SCRIPTS)}")
                else:
                    logger.info(f"所有抓图脚本执行成功: {bin_location}")

                if ((index + 1) < total):
                    logger.info(f"收到机器人结束状态: {bin_location}")

                    # 只有在收到end状态后才调用继续任务接口
                    continue_result = await continue_inventory_task()
                    logger.info(f"继续任务接口调用结果: {continue_result}")
                    result["continueResult"] = continue_result

            else:
                # 正常情况下不会执行到这里，除非wait_for_robot_status返回了非end状态
                logger.warning(f"未收到预期的结束状态，当前状态: {ctu_status}")

        except asyncio.TimeoutError as e:
            logger.error(f"等待机器人结束状态超时: {str(e)}")
            result["error"] = "等待机器人结束状态超时"
            raise

        # 2-4. 执行完整的盘点任务流程：抓图 -> 计算 -> 发送
        inventory_service = get_inventory_service()
        inventory_result = await inventory_service.process_inventory_task(
            task_no=task_no,
            bin_location=bin_location
        )

        result["imageData"] = inventory_result.get("imageData")
        result["captureTime"] = inventory_result.get("captureTime")
        result["computeResult"] = inventory_result.get("computeResult")
        result["computeTime"] = inventory_result.get("computeTime")

        # 更新任务状态中的图片和计算结果（供前端获取）
        if task_no in inventory_task_bins:
            for bin_status in inventory_task_bins[task_no]:
                if bin_status.bin_location == bin_location:
                    bin_status.image_data = inventory_result.get("imageData")
                    bin_status.compute_result = inventory_result.get(
                        "computeResult")
                    bin_status.capture_time = inventory_result.get(
                        "captureTime")
                    bin_status.compute_time = inventory_result.get(
                        "computeTime")
                    break

        # 存储详细结果
        if task_no not in inventory_task_details:
            inventory_task_details[task_no] = {}

        inventory_task_details[task_no][bin_location] = {
            "image_data": inventory_result.get("imageData"),
            "compute_result": inventory_result.get("computeResult"),
            "capture_time": inventory_result.get("captureTime"),
            "compute_time": inventory_result.get("computeTime"),
            "updated_at": datetime.now().isoformat()
        }

        result["status"] = "success"
        result["endTime"] = datetime.now().isoformat()

    except Exception as e:
        result["status"] = "failed"
        result["endTime"] = datetime.now().isoformat()
        logger.error(f"处理储位失败 {bin_location}: {str(e)}")

        # 记录错误但继续处理下一个储位（根据业务需求决定是否中断）
        # 可以发送错误通知到前端
        # try:
        #     async with APIClient(SERVICE_CONFIG["notification_service"]) as client:
        #         error_payload = {
        #             "taskNo": task_no,
        #             "binLocation": bin_location,
        #             "error": str(e),
        #             "timestamp": datetime.now().isoformat(),
        #             "messageType": "ERROR"
        #         }
        #         await client.post("/api/notification/error", json=error_payload)
        # except:
        #     pass

    return result


async def execute_inventory_workflow(task_no: str, bin_locations: List[str]):
    """执行完整的盘点工作流"""
    logger.info(f"开始执行盘点工作流: {task_no}, 共 {len(bin_locations)} 个储位")

    # 初始化任务状态
    task_status = TaskStatus(
        task_no=task_no,
        status="init",
        current_step=0,
        total_steps=len(bin_locations),
        start_time=datetime.now().isoformat()
    )

    # 按任务编号存储任务状态
    inventory_tasks[task_no] = task_status

    # 初始化每个储位的状态
    bin_statuses = [
        BinLocationStatus(
            bin_location=location,
            status="pending",
            sequence=index + 1
        )
        for index, location in enumerate(bin_locations)
    ]
    inventory_task_bins[task_no] = bin_statuses

    # 更新任务状态为运行中
    inventory_tasks[task_no].status = "running"
    inventory_tasks[task_no].current_step = 0

    # 整体下发盘点任务
    method = "start"
    await update_robot_status(method)

    submit_result = await submit_inventory_task(task_no, bin_locations)

    try:
        # 循环处理每个储位
        for i, bin_location in enumerate(bin_locations):
            logger.info(f"开始处理储位 {i+1}/{len(bin_locations)}: {bin_location}")

            # 更新当前步骤
            inventory_tasks[task_no].current_step = i + 1

            # 更新储位状态为运行中
            if task_no in inventory_task_bins:
                for bin_status in inventory_task_bins[task_no]:
                    if bin_status.bin_location == bin_location:
                        bin_status.status = "running"
                        break

            # 处理单个储位
            result = await process_single_bin_location(
                task_no=task_no,
                bin_location=bin_location,
                index=i,
                total=len(bin_locations)
            )

            # 保存结果
            if task_no in inventory_task_bins:
                for bin_status in inventory_task_bins[task_no]:
                    if bin_status.bin_location == bin_location:
                        if result["status"] == "success":
                            bin_status.status = "completed"
                        else:
                            bin_status.status = "failed"
                        break

            if result["status"] != "success":
                inventory_tasks[task_no].status = "failed"
                raise Exception("储位处理失败，终止任务")

        # 更新任务状态为完成
        inventory_tasks[task_no].status = "completed"
        inventory_tasks[task_no].current_step = len(bin_locations)
        inventory_tasks[task_no].end_time = datetime.now().isoformat()

        logger.info(f"盘点任务完成: {task_no}, 成功处理 {len(bin_locations)} 个储位")

        # 记录任务完成
        log_operation(
            operation_type="inventory",
            action="盘点任务完成",
            target=task_no,
            status="completed",
            details={
                "task_no": task_no,
                "bin_locations": bin_locations,
                "completed_count": len(bin_locations),
                "xlsx_file": f"./output/history_data/{task_no}.xlsx"
            }
        )

        # 发送任务完成通知
        # try:
        #     async with APIClient(SERVICE_CONFIG["notification_service"]) as client:
        #         completion_payload = {
        #             "taskNo": task_no,
        #             "status": "COMPLETED",
        #             "totalBins": len(bin_locations),
        #             "successfulBins": sum(1 for r in inventory_tasks[task_no].results
        #                                   if r.get("status") == "completed"),
        #             "failedBins": sum(1 for r in inventory_tasks[task_no].results
        #                               if r.get("status") == "failed"),
        #             "completionTime": datetime.now().isoformat(),
        #             "messageType": "TASK_COMPLETED"
        #         }
        #         await client.post("/api/notification/task-complete", json=completion_payload)
        # except Exception as e:
        #     logger.warning(f"发送任务完成通知失败: {str(e)}")

    except Exception as e:
        # 任务执行过程中出现异常
        if task_no in inventory_tasks:
            inventory_tasks[task_no].status = "failed"
            inventory_tasks[task_no].end_time = datetime.now().isoformat()

        # 记录任务失败
        log_operation(
            operation_type="inventory",
            action="盘点任务失败",
            target=task_no,
            status="failed",
            details={
                "task_no": task_no,
                "error": str(e),
                "failed_at_step": inventory_tasks[task_no].current_step if task_no in inventory_tasks else 0
            }
        )

        logger.error(f"盘点任务失败: {task_no}, 错误: {str(e)}")

        # 发送任务失败通知
        # try:
        #     async with APIClient(SERVICE_CONFIG["notification_service"]) as client:
        #         error_payload = {
        #             "taskNo": task_no,
        #             "status": "FAILED",
        #             "error": str(e),
        #             "failedAtBin": inventory_tasks[task_no].current_bin,
        #             "completedBins": len(inventory_tasks[task_no].results),
        #             "timestamp": datetime.now().isoformat(),
        #             "messageType": "TASK_FAILED"
        #         }
        #         await client.post("/api/notification/task-error", json=error_payload)
        # except Exception as e2:
        #     logger.error(f"发送任务失败通知失败: {str(e2)}")


######################################### 盘点任务接口 #########################################


@app.post("/api/inventory/start-inventory")
async def start_inventory(request: Request, background_tasks: BackgroundTasks):
    """启动盘点任务，接收任务编号和储位名称列表"""
    try:
        data = await request.json()
        task_no = data.get("taskNo")
        bin_locations = data.get("binLocations", [])

        if not task_no or not bin_locations:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="任务编号和储位名称列表不能为空"
            )

        logger.info(f"启动盘点任务: {task_no}, 包含 {len(bin_locations)} 个储位")

        # 检查任务是否已存在
        if task_no in inventory_tasks:
            existing_task = inventory_tasks[task_no]
            if existing_task.status in ["running", "init"]:
                return JSONResponse(
                    status_code=status.HTTP_200_OK,
                    content={
                        "code": 200,
                        "message": "任务已在执行中",
                        "data": {
                            "taskNo": existing_task.task_no,
                            "status": existing_task.status,
                        }
                    }
                )

        # 记录盘点任务启动
        auth_token = request.headers.get("authToken")
        user_info = await get_user_info_from_token(auth_token) if auth_token else {}

        log_operation(
            operation_type="inventory",
            action="启动盘点任务",
            user_id=user_info.get("userId"),
            user_name=user_info.get("userName"),
            target=task_no,
            status="running",
            details={
                "task_no": task_no,
                "bin_locations": bin_locations,
                "bin_count": len(bin_locations)
            }
        )

        # 在后台异步执行盘点任务
        background_tasks.add_task(
            execute_inventory_workflow,
            task_no=task_no,
            bin_locations=bin_locations
        )

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "code": 200,
                "message": "盘点任务已启动",
                "data": {
                    "taskNo": task_no,
                    "bin_locations": bin_locations
                }
            }
        )

    except Exception as e:
        logger.error(f"启动盘点任务失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"启动盘点任务失败: {str(e)}"
        )


@app.get("/api/inventory/progress")
async def get_inventory_progress(taskNo: str):
    """获取盘点任务进度"""
    try:
        if taskNo not in inventory_tasks:
            return JSONResponse(
                status_code=status.HTTP_404_NOT_FOUND,
                content={
                    "code": 404,
                    "message": "任务不存在",
                    "data": None
                }
            )

        task_status = inventory_tasks[taskNo]
        bin_statuses = inventory_task_bins.get(taskNo, [])

        # 计算进度百分比
        completed_count = sum(
            1 for bin in bin_statuses if bin.status == "completed")
        progress_percentage = (
            completed_count / task_status.total_steps * 100) if task_status.total_steps > 0 else 0

        progress_data = InventoryTaskProgress(
            task_no=task_status.task_no,
            status=task_status.status,
            current_step=task_status.current_step,
            total_steps=task_status.total_steps,
            progress_percentage=round(progress_percentage, 2),
            bin_locations=bin_statuses,
            start_time=task_status.start_time,
            end_time=task_status.end_time
        )

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "code": 200,
                "message": "获取进度成功",
                "data": progress_data.dict()
            }
        )
    except Exception as e:
        logger.error(f"获取任务进度失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取任务进度失败: {str(e)}"
        )


@app.get("/api/inventory/image")
async def get_inventory_image(
    taskNo: str,
    binLocation: str,
    cameraType: str,
    filename: str,
    source: str = "output"  # 新增参数：source可以是"output"或"capture_img"
):
    """
    获取盘点任务中的图片

    Args:
        taskNo: 任务编号
        binLocation: 储位名称
        cameraType: 相机类型
        filename: 文件名
        source: 图片源目录，可以是"output"或"capture_img"，默认为"output"
    """
    try:
        project_root = Path(__file__).parent.parent.parent

        # 根据source参数选择不同的基础目录
        if source == "capture_img":
            # 从capture_img目录获取图片
            image_path = project_root / "capture_img" / \
                taskNo / binLocation / cameraType / filename
        else:
            # 从output目录获取图片（默认）
            image_path = project_root / "output" / \
                taskNo / binLocation / cameraType / filename

        if not image_path.exists():
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"图片不存在: {filename} (路径: {image_path})"
            )

        # 读取图片文件
        with open(image_path, "rb") as f:
            image_data = f.read()

        # 根据文件扩展名确定媒体类型
        media_type = "image/jpeg"
        if filename.endswith(".png"):
            media_type = "image/png"
        elif filename.endswith(".bmp"):
            media_type = "image/bmp"

        return Response(content=image_data, media_type=media_type)

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取图片失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取图片失败: {str(e)}"
        )


@app.get("/api/inventory/task-detail")
async def get_task_detail(taskNo: str, binLocation: str):
    """
    获取任务的详细信息，包括图片和计算结果

    Args:
        taskNo: 任务编号
        binLocation: 储位名称
    """
    try:
        # 从任务详情中获取（使用存储函数）
        storage = get_task_state_storage()
        task_details = storage["details"]

        if taskNo in task_details and binLocation in task_details[taskNo]:
            detail = task_details[taskNo][binLocation]
            return JSONResponse(
                status_code=status.HTTP_200_OK,
                content={
                    "code": 200,
                    "message": "获取任务详情成功",
                    "data": detail
                }
            )
        else:
            return JSONResponse(
                status_code=status.HTTP_404_NOT_FOUND,
                content={
                    "code": 404,
                    "message": "任务详情不存在",
                    "data": None
                }
            )

    except Exception as e:
        logger.error(f"获取任务详情失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取任务详情失败: {str(e)}"
        )


@app.post("/api/inventory/scan-and-recognize")
async def scan_and_recognize(request: ScanAndRecognizeRequest = Body(...)):
    """
    扫码+识别接口
    同时执行Detect模块和Barcode模块的识别

    Args:
        request: ScanAndRecognizeRequest对象，包含taskNo、binLocation、pile_id和code_type
    """
    try:
        project_root = Path(__file__).parent.parent.parent

        # 构建图片路径: taskNo/binLocation/3d_camera/
        image_path = f"{request.taskNo}/{request.binLocation}/3d_camera/"
        image_dir = project_root / "capture_img" / image_path

        recognition_time = datetime.now().isoformat()
        results = {
            "taskNo": request.taskNo,
            "binLocation": request.binLocation,
            "recognition_time": recognition_time,
            "detect_result": None,
            "barcode_result": None,
            "photos": []  # 添加照片路径数组
        }

        # 查找并返回照片路径（无论检测模块是否可用）
        try:
            photos = []
            image_extensions = ['.jpg', '.jpeg', '.png', '.bmp']

            # 优先查找主图 (3d_camera 目录)
            main_images = []
            if image_dir.exists():
                for ext in image_extensions:
                    # 查找常见的文件名
                    for common_name in ['main', 'MAIN', 'raw', 'RAW', 'image', 'IMAGE']:
                        main_file = image_dir / f"{common_name}{ext}"
                        if main_file.exists():
                            # 构建相对于 capture_img 的路径
                            relative_path = f"{request.taskNo}/{request.binLocation}/3d_camera/{common_name.upper()}{ext}"
                            main_images.append(f"/{relative_path}")
                            logger.info(f"找到主图: {relative_path}")
                            break
                    if main_images:
                        break

                # 如果没有找到常见文件名，查找第一张图片
                if not main_images:
                    for ext in image_extensions:
                        image_files = list(image_dir.glob(f"*{ext}"))
                        if image_files:
                            main_file = image_files[0]
                            relative_path = f"{request.taskNo}/{request.binLocation}/3d_camera/{main_file.name}"
                            photos.append(f"/{relative_path}")
                            logger.info(f"找到主图: {relative_path}")
                            break

                photos.extend(main_images)

                # 查找深度图 (depth 目录)
                depth_dir = project_root / "capture_img" / \
                    f"{request.taskNo}/{request.binLocation}/depth"
                if depth_dir.exists():
                    for ext in image_extensions:
                        # 查找常见的文件名
                        for common_name in ['color', 'COLOR', 'depth', 'DEPTH']:
                            depth_file = depth_dir / f"{common_name}{ext}"
                            if depth_file.exists():
                                relative_path = f"{request.taskNo}/{request.binLocation}/depth/{common_name.upper()}{ext}"
                                photos.append(f"/{relative_path}")
                                logger.info(f"找到深度图: {relative_path}")
                                break
                        if len(photos) > len(main_images):  # 已经找到深度图
                            break

                        # 如果没有找到常见文件名，查找第一张图片
                        depth_files = list(depth_dir.glob(f"*{ext}"))
                        if depth_files and len(photos) == len(main_images):
                            depth_file = depth_files[0]
                            relative_path = f"{request.taskNo}/{request.binLocation}/depth/{depth_file.name}"
                            photos.append(f"/{relative_path}")
                            logger.info(f"找到深度图: {relative_path}")
                            break

            results["photos"] = photos
            logger.info(f"共找到 {len(photos)} 张图片")

        except Exception as e:
            logger.error(f"查找照片路径失败: {str(e)}", exc_info=True)
            results["photos"] = []  # 失败时返回空数组，不中断流程

        # 如果检测模块可用，执行识别
        if DETECT_MODULE_AVAILABLE:
            # 严格检查 capture_img 下的路径是否存在
            if not image_dir.exists() or not image_dir.is_dir():
                logger.error(f"图片目录不存在: {image_dir}")
                # 即使目录不存在，也返回照片路径（为空）
                return JSONResponse(
                    status_code=status.HTTP_200_OK,
                    content={
                        "code": 200,
                        "message": "图片目录不存在，但返回了照片路径信息",
                        "data": results
                    }
                )

            # 1. 执行Detect模块
            try:
                logger.info(
                    f"开始执行Detect模块识别: {request.taskNo}/{request.binLocation}")

                # 查找目录中的图片文件
                image_extensions = ['.jpg', '.jpeg', '.png', '.bmp']
                image_files = []

                # 优先查找常见文件名
                common_names = ['main', 'raw', 'image', 'img', 'photo']
                for name in common_names:
                    for ext in image_extensions:
                        common_file = image_dir / f"{name}{ext}"
                        if common_file.exists():
                            image_files.append(common_file)
                            break
                    if image_files:
                        break

                # 如果没找到，查找所有图片
                if not image_files:
                    for ext in image_extensions:
                        image_files.extend(list(image_dir.glob(f"*{ext}")))
                        if image_files:
                            break

                if not image_files:
                    raise HTTPException(
                        status_code=status.HTTP_404_NOT_FOUND,
                        detail=f"在路径 {image_dir} 中未找到图片文件"
                    )

                image_path_for_detect = str(image_files[0])

                # 从输入路径加载深度图 depth.jpg
                depth_image_path = image_dir / "depth.jpg"
                depth_image_path_for_detect = None
                if depth_image_path.exists():
                    depth_image_path_for_detect = str(depth_image_path)
                    logger.info(f"找到深度图文件: {depth_image_path_for_detect}")
                else:
                    logger.warning(f"深度图文件不存在: {depth_image_path}，将不使用深度图")

                # 构建debug输出目录: debug/{taskNo}/{binLocation}/
                debug_output_dir = project_root / "debug" / request.taskNo / request.binLocation
                debug_output_dir.mkdir(parents=True, exist_ok=True)

                # 创建日志文件处理器，将算法日志保存到debug目录
                log_file_path = debug_output_dir / \
                    f"debug_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
                file_handler = logging.FileHandler(
                    str(log_file_path), encoding='utf-8')
                file_handler.setLevel(logging.DEBUG)
                file_formatter = logging.Formatter(
                    '%(asctime)s - %(name)s - %(levelname)s - %(message)s',
                    datefmt='%Y-%m-%d %H:%M:%S'
                )
                file_handler.setFormatter(file_formatter)

                # 添加文件处理器到logger（临时添加，执行完后移除）
                logger.addHandler(file_handler)

                try:
                    logger.info(f"Debug输出目录: {debug_output_dir}")
                    logger.info(f"Debug日志文件: {log_file_path}")
                    logger.info(f"使用主图片: {image_path_for_detect}")
                    if depth_image_path_for_detect:
                        logger.info(f"使用深度图: {depth_image_path_for_detect}")

                    # 重定向stdout和stderr到日志文件，捕获count_boxes的print输出
                    import io
                    from contextlib import redirect_stdout, redirect_stderr

                    # 创建一个类，将写入操作转换为日志记录
                    class LogWriter:
                        def __init__(self, logger_instance, level=logging.INFO):
                            self.logger = logger_instance
                            self.level = level
                            self.buffer = ""

                        def write(self, message):
                            if message.strip():  # 忽略空行
                                self.buffer += message
                                # 检查是否有完整的行
                                while '\n' in self.buffer:
                                    line, self.buffer = self.buffer.split(
                                        '\n', 1)
                                    if line.strip():
                                        self.logger.log(
                                            self.level, line.strip())

                        def flush(self):
                            if self.buffer.strip():
                                self.logger.log(
                                    self.level, self.buffer.strip())
                                self.buffer = ""

                    log_writer = LogWriter(logger, logging.INFO)

                    # 执行count_boxes，同时捕获其print输出
                    with redirect_stdout(log_writer), redirect_stderr(log_writer):
                        total_count = count_boxes(
                            image_path=image_path_for_detect,
                            pile_id=request.pile_id,
                            depth_image_path=depth_image_path_for_detect,
                            enable_debug=False,
                            enable_visualization=False,
                            output_dir=str(debug_output_dir)
                        )

                    logger.info(f"Detect模块识别完成，箱数: {total_count}")

                    results["detect_result"] = {
                        "image_path": image_path_for_detect,
                        "pile_id": request.pile_id,
                        "total_count": total_count,
                        "status": "success"
                    }

                except Exception as e:
                    logger.error(f"Detect模块识别失败: {str(e)}", exc_info=True)
                    results["detect_result"] = {
                        "status": "failed",
                        "error": str(e)
                    }
                finally:
                    # 移除文件处理器，避免日志重复输出
                    logger.removeHandler(file_handler)
                    file_handler.close()

            except HTTPException:
                raise
            except Exception as e:
                logger.error(f"Detect模块初始化失败: {str(e)}", exc_info=True)
                results["detect_result"] = {
                    "status": "failed",
                    "error": str(e)
                }

        # 2. 执行Barcode模块（仅在开关开启时执行）
        if ENABLE_BARCODE and BARCODE_MODULE_AVAILABLE:
            try:
                logger.info(
                    f"开始执行Barcode模块识别: {request.taskNo}/{request.binLocation}")

                recognizer = BarcodeRecognizer(code_type=request.code_type)
                barcode_results = recognizer.process_folder(
                    input_dir=str(image_dir))

                results["barcode_result"] = {
                    "image_path": str(image_dir),
                    "code_type": request.code_type,
                    "results": barcode_results,
                    "total_images": len(barcode_results),
                    "successful": sum(1 for r in barcode_results if r.get("output")),
                    "failed": sum(1 for r in barcode_results if r.get("error") and not r.get("output")),
                    "status": "success"
                }
                logger.info(
                    f"Barcode模块识别完成，成功: {results['barcode_result']['successful']}/{results['barcode_result']['total_images']}")

            except Exception as e:
                logger.error(f"Barcode模块识别失败: {str(e)}")
                results["barcode_result"] = {
                    "status": "failed",
                    "error": str(e)
                }
        else:
            logger.info("Barcode模块已禁用，跳过识别")
            results["barcode_result"] = {
                "status": "disabled",
                "message": "条形码功能已禁用（ENABLE_BARCODE=false）"
            }

        # 3. 更新任务状态中的识别结果
        storage = get_task_state_storage()
        inventory_task_bins = storage["bins"]
        inventory_task_details = storage["details"]

        if request.taskNo in inventory_task_bins:
            for bin_status in inventory_task_bins[request.taskNo]:
                if bin_status.bin_location == request.binLocation:
                    bin_status.detect_result = results["detect_result"]
                    bin_status.barcode_result = results["barcode_result"]
                    bin_status.recognition_time = recognition_time
                    break

        # 4. 存储到任务详情中
        if request.taskNo not in inventory_task_details:
            inventory_task_details[request.taskNo] = {}

        if request.binLocation not in inventory_task_details[request.taskNo]:
            inventory_task_details[request.taskNo][request.binLocation] = {}

        inventory_task_details[request.taskNo][request.binLocation]["recognition"] = results

        # 5. 将识别结果保存为ini格式文件到原始图路径
        try:
            import configparser

            # 创建ConfigParser对象
            config = configparser.ConfigParser()

            # 添加基本信息节
            config.add_section('BasicInfo')
            config.set('BasicInfo', 'taskNo', request.taskNo)
            config.set('BasicInfo', 'binLocation', request.binLocation)
            config.set('BasicInfo', 'recognition_time', recognition_time)

            # 添加原始图路径
            config.add_section('ImagePaths')
            config.set('ImagePaths', 'original_image_dir', str(image_dir))
            if results.get("detect_result") and results["detect_result"].get("image_path"):
                config.set('ImagePaths', 'detect_image_path',
                           results["detect_result"]["image_path"])

            # 添加Detect模块结果
            if results.get("detect_result"):
                config.add_section('DetectResult')
                detect_result = results["detect_result"]
                if detect_result.get("status"):
                    config.set('DetectResult', 'status',
                               detect_result["status"])
                if detect_result.get("total_count") is not None:
                    config.set('DetectResult', 'total_count',
                               str(detect_result["total_count"]))
                if detect_result.get("pile_id") is not None:
                    config.set('DetectResult', 'pile_id',
                               str(detect_result["pile_id"]))
                if detect_result.get("error"):
                    config.set('DetectResult', 'error', detect_result["error"])

            # 添加Barcode模块结果
            if results.get("barcode_result"):
                config.add_section('BarcodeResult')
                barcode_result = results["barcode_result"]
                if barcode_result.get("status"):
                    config.set('BarcodeResult', 'status',
                               barcode_result["status"])
                if barcode_result.get("code_type"):
                    config.set('BarcodeResult', 'code_type',
                               barcode_result["code_type"])
                if barcode_result.get("total_images") is not None:
                    config.set('BarcodeResult', 'total_images',
                               str(barcode_result["total_images"]))
                if barcode_result.get("successful") is not None:
                    config.set('BarcodeResult', 'successful',
                               str(barcode_result["successful"]))
                if barcode_result.get("failed") is not None:
                    config.set('BarcodeResult', 'failed',
                               str(barcode_result["failed"]))
                if barcode_result.get("error"):
                    config.set('BarcodeResult', 'error',
                               barcode_result["error"])
                if barcode_result.get("message"):
                    config.set('BarcodeResult', 'message',
                               barcode_result["message"])

            # 保存ini文件到原始图路径
            ini_file_path = image_dir / "recognition_result.ini"
            with open(ini_file_path, 'w', encoding='utf-8') as configfile:
                config.write(configfile)

            logger.info(f"识别结果已保存到ini文件: {ini_file_path}")
        except Exception as e:
            logger.error(f"保存识别结果到ini文件失败: {str(e)}", exc_info=True)
            # 不中断流程，继续返回结果

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "code": 200,
                "message": "扫码+识别执行完成",
                "data": results
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"扫码+识别失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"扫码+识别失败: {str(e)}"
        )


@app.get("/api/inventory/recognition-result")
async def get_recognition_result(taskNo: str, binLocation: str):
    """
    读取识别结果接口
    获取指定任务和库位的Detect和Barcode识别结果

    Args:
        taskNo: 任务编号
        binLocation: 库位号
    """
    try:
        # 获取任务状态存储
        storage = get_task_state_storage()
        inventory_task_bins = storage["bins"]
        inventory_task_details = storage["details"]

        result_data = {
            "taskNo": taskNo,
            "binLocation": binLocation,
            "detect_result": None,
            "barcode_result": None,
            "recognition_time": None
        }

        # 从任务状态中获取（使用存储函数）
        storage = get_task_state_storage()
        task_bins = storage["bins"]
        task_details = storage["details"]

        if taskNo in task_bins:
            for bin_status in task_bins[taskNo]:
                if bin_status.bin_location == binLocation:
                    result_data["detect_result"] = bin_status.detect_result
                    result_data["barcode_result"] = bin_status.barcode_result
                    result_data["recognition_time"] = bin_status.recognition_time
                    break

        # 如果状态中没有，尝试从任务详情中获取
        if not result_data["detect_result"] and taskNo in task_details:
            if binLocation in task_details[taskNo]:
                recognition_data = task_details[taskNo][binLocation].get(
                    "recognition")
                if recognition_data:
                    result_data["detect_result"] = recognition_data.get(
                        "detect_result")
                    result_data["barcode_result"] = recognition_data.get(
                        "barcode_result")
                    result_data["recognition_time"] = recognition_data.get(
                        "recognition_time")

        # 检查是否有结果
        if not result_data["detect_result"] and not result_data["barcode_result"]:
            return JSONResponse(
                status_code=status.HTTP_404_NOT_FOUND,
                content={
                    "code": 404,
                    "message": "识别结果不存在",
                    "data": None
                }
            )

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "code": 200,
                "message": "获取识别结果成功",
                "data": result_data
            }
        )

    except Exception as e:
        error_msg = f"获取识别结果失败: {str(e)}"
        logger.error(error_msg, exc_info=True)  # exc_info=True 会输出完整的堆栈信息
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=error_msg
        )


######################################### LMS #########################################


@app.get("/auth/token")
async def auth_token(token: str):
    """处理前端获取用户信息请求，调用LMS的authToken接口"""
    try:
        # 调用LMS的authToken接口
        lms_auth_url = f"{LMS_BASE_URL}/auth/token?token={token}"
        response = requests.get(lms_auth_url)

        if response.status_code == 200:
            return response.json()
        else:
            raise HTTPException(
                status_code=response.status_code,
                detail=f"LMS获取用户信息失败: {response.text}"
            )
    except Exception as e:
        logger.error(f"获取用户信息请求失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="获取用户信息请求处理失败"
        )


@app.get("/lms/getLmsBin")
async def get_lms_bin(authToken: str):
    """获取库位信息，调用LMS的getLmsBin接口"""
    try:
        # 调用LMS的getLmsBin接口
        lms_bin_url = f"{LMS_BASE_URL}/third/api/v1/lmsToRcsService/getLmsBin"
        headers = {
            "authToken": authToken
        }
        response = requests.get(lms_bin_url, headers=headers)

        if response.status_code == 200:
            # 关键修复：处理LMS返回的压缩编码字符串
            try:
                uncompressed_data = custom_utils.decompress_and_decode(
                    response.text)

                logger.info("成功解压缩并解析库位数据")
                return JSONResponse(uncompressed_data)
            except Exception as e:
                logger.error(f"解压缩库位数据失败: {str(e)}")
                raise HTTPException(
                    status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    detail="库位数据解压缩失败"
                )
        else:
            raise HTTPException(
                status_code=response.status_code,
                detail=f"LMS获取库位信息失败: {response.text}"
            )
    except Exception as e:
        logger.error(f"获取库位信息请求失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="获取库位信息请求处理失败"
        )


@app.get("/lms/getCountTasks")
async def get_count_tasks(authToken: str):
    """获取盘点任务，调用LMS的getCountTasks接口"""
    try:
        logger.info(f"收到获取盘点任务请求，authToken: {authToken[:20]}...")

        lms_tasks_url = f"{LMS_BASE_URL}/third/api/v1/lmsToRcsService/getCountTasks"
        logger.info(f"准备调用LMS接口: {lms_tasks_url}")

        headers = {"authToken": authToken}
        logger.info("发送请求到LMS服务...")
        response = requests.get(lms_tasks_url, headers=headers, timeout=30)
        logger.info(f"LMS响应状态码: {response.status_code}")

        if response.status_code == 200:
            # 关键修复：处理LMS返回的压缩编码字符串
            try:
                uncompressed_data = custom_utils.decompress_and_decode(
                    response.text)

                logger.info("成功解压缩并解析盘点任务数据")
                return JSONResponse(uncompressed_data)
            except Exception as e:
                logger.error(f"解压缩盘点任务数据失败: {str(e)}")
                raise HTTPException(
                    status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    detail="盘点任务数据解压缩失败"
                )
        else:
            logger.error(
                f"LMS获取盘点任务失败: {response.status_code} - {response.text}")
            raise HTTPException(
                status_code=response.status_code,
                detail=f"LMS获取盘点任务失败: {response.text}"
            )
    except requests.exceptions.Timeout:
        logger.error("LMS服务请求超时")
        raise HTTPException(
            status_code=status.HTTP_504_GATEWAY_TIMEOUT,
            detail="LMS服务响应超时"
        )
    except requests.exceptions.ConnectionError:
        logger.error("无法连接到LMS服务")
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="无法连接到LMS服务"
        )
    except Exception as e:
        logger.error(f"获取盘点任务请求失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="获取盘点任务请求处理失败"
        )


@app.post("/lms/setTaskResults")
async def set_task_results(request: Request):
    """提交盘点任务结果，调用LMS的setTaskResults接口"""
    try:
        # 1. 从请求头获取authToken
        auth_token = request.headers.get('authToken')
        if not auth_token:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Unauthorized"
            )

        # 2. 从请求体获取JSON数据（前端发送的是标准JSON）
        data = await request.json()
        encoded_data = custom_utils.compress_and_encode(data)

        # 6. 调用LMS接口（使用压缩后的数据）
        lms_results_url = f"{LMS_BASE_URL}/third/api/v1/RcsToLmsService/setTaskResults"
        headers = {
            "authToken": auth_token,  # 传递给LMS的认证令牌
            "Content-Type": "text/plain"  # 关键：必须是text/plain
        }

        # 发送压缩后的base64字符串
        response = requests.post(
            lms_results_url, data=encoded_data, headers=headers)

        if response.status_code == 200:
            return {"success": True, "message": "盘点结果已提交"}
        else:
            raise HTTPException(
                status_code=response.status_code,
                detail=f"LMS提交盘点结果失败: {response.text}"
            )
    except Exception as e:
        logger.error(f"提交盘点结果请求失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="提交盘点结果请求处理失败"
        )


######################################### RCS #########################################

@app.post("/api/robot/reporter/task")
async def task_status(request: Request):
    try:
        # 获取请求数据
        request_data = await request.json()

        logger.info("反馈任务状态")
        logger.info(
            f"请求数据: {json.dumps(request_data, indent=2, ensure_ascii=False)}")

        # 提取任务信息
        robot_task_code = request_data.get("robotTaskCode")
        single_robot_code = request_data.get("singleRobotCode")
        extra = request_data.get("extra", "")

        # 解析extra字段
        if extra:
            try:
                extra_list = json.loads(extra)
                if isinstance(extra_list, list):
                    for item in extra_list:
                        method = item.get("method", "")
                        logger.info(f"处理method: {method}")
                        await update_robot_status(method, item)

                        if method == "start":
                            logger.info("任务开始")

                        elif method == "outbin":
                            logger.info("走出储位")

                        elif method == "end":
                            logger.info("任务完成")

                        # 根据不同的method更新您的任务状态...
            except json.JSONDecodeError:
                logger.error(f"无法解析extra字段: {extra}")

        # 返回响应
        return {
            "code": "SUCCESS",
            "message": "成功",
            "data": {
                "robotTaskCode": "ctu001"
            }
        }

    except Exception as e:
        logger.error(f"处理状态反馈失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"处理状态反馈失败: {str(e)}")


# 前端日志收集接口
@app.post("/api/log/frontend")
async def collect_frontend_log(request: FrontendLogRequest = Body(...)):
    """
    收集前端日志并保存到 debug 目录

    Args:
        request: 前端日志请求对象
    """
    try:
        # 创建前端日志文件路径（按日期命名）
        frontend_log_filename = _debug_log_dir / \
            f"frontend_{datetime.now().strftime('%Y%m%d')}.log"

        # 创建前端日志记录器
        frontend_logger = logging.getLogger("frontend")
        frontend_logger.setLevel(logging.DEBUG)

        # 如果还没有文件处理器，添加一个
        if not any(isinstance(h, logging.FileHandler) and h.baseFilename == str(frontend_log_filename)
                   for h in frontend_logger.handlers):
            file_handler = logging.FileHandler(
                str(frontend_log_filename), encoding='utf-8')
            file_handler.setLevel(logging.DEBUG)
            formatter = logging.Formatter(
                '%(asctime)s - [FRONTEND] - %(levelname)s - %(message)s',
                datefmt='%Y-%m-%d %H:%M:%S'
            )
            file_handler.setFormatter(formatter)
            frontend_logger.addHandler(file_handler)

        # 记录日志
        timestamp = request.timestamp or datetime.now().isoformat()
        source = request.source or "unknown"
        extra_info = request.extra or {}

        log_message = f"[{source}] {request.message}"
        if extra_info:
            log_message += f" | Extra: {json.dumps(extra_info, ensure_ascii=False)}"

        # 根据日志级别记录
        log_level = request.level.lower()
        if log_level == "error":
            frontend_logger.error(log_message)
        elif log_level == "warn":
            frontend_logger.warning(log_message)
        elif log_level == "info":
            frontend_logger.info(log_message)
        else:
            frontend_logger.debug(log_message)

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "code": 200,
                "message": "日志已保存",
                "data": {
                    "log_file": str(frontend_log_filename)
                }
            }
        )

    except Exception as e:
        logger.error(f"保存前端日志失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"保存前端日志失败: {str(e)}"
        )


######################################### 用户管理接口 #########################################

@app.get("/lms/getUsers")
async def get_users(request: Request):
    """获取所有用户信息，调用LMS的getUsers接口"""
    try:
        # 从查询参数或请求头获取authToken
        auth_token = request.query_params.get(
            'authToken') or request.headers.get('authToken')

        if not auth_token:
            logger.error("未提供认证令牌")
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="未提供认证令牌"
            )

        logger.info(f"接收到的authToken: {auth_token[:20]}...")

        # 调用LMS的getUsers接口
        lms_users_url = f"{LMS_BASE_URL}/third/api/v1/userManagement/getUsers"
        headers = {
            "authToken": auth_token
        }

        logger.info(f"调用LMS用户接口: {lms_users_url}")

        try:
            response = requests.get(lms_users_url, headers=headers, timeout=10)
            logger.info(f"LMS响应状态码: {response.status_code}")

            # 检查响应状态
            if response.status_code == 200:
                # 尝试解析JSON响应
                try:
                    result = response.json()
                    logger.info(f"成功获取 {len(result.get('data', []))} 条用户数据")
                    return result
                except json.JSONDecodeError as e:
                    logger.error(f"LMS返回的不是有效JSON: {response.text[:200]}")
                    raise HTTPException(
                        status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                        detail=f"LMS返回了无效的JSON数据: {str(e)}"
                    )
            else:
                # 记录详细的错误信息
                error_detail = response.text[:500]  # 只取前500字符避免日志过大
                logger.error(f"LMS接口错误 {response.status_code}: {error_detail}")

                if response.status_code == 401:
                    return JSONResponse(
                        status_code=status.HTTP_401_UNAUTHORIZED,
                        content={
                            "code": 401,
                            "message": "认证失败，请重新登录",
                            "data": None
                        }
                    )
                elif response.status_code == 403:
                    return JSONResponse(
                        status_code=status.HTTP_403_FORBIDDEN,
                        content={
                            "code": 403,
                            "message": "权限不足，只有管理员可以查看用户列表",
                            "data": None
                        }
                    )
                else:
                    return JSONResponse(
                        status_code=response.status_code,
                        content={
                            "code": response.status_code,
                            "message": f"LMS获取用户信息失败: {response.text[:200]}",
                            "data": None
                        }
                    )

        except requests.exceptions.ConnectionError:
            logger.error(f"无法连接到LMS服务: {LMS_BASE_URL}")
            return JSONResponse(
                status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
                content={
                    "code": 503,
                    "message": f"无法连接到LMS服务，请确保LMS服务正在运行（{LMS_BASE_URL}）",
                    "data": None
                }
            )
        except requests.exceptions.Timeout:
            logger.error(f"连接LMS服务超时")
            return JSONResponse(
                status_code=status.HTTP_504_GATEWAY_TIMEOUT,
                content={
                    "code": 504,
                    "message": "LMS服务响应超时",
                    "data": None
                }
            )

    except HTTPException as he:
        # 重新抛出 HTTPException
        raise he
    except Exception as e:
        logger.error(f"获取用户信息请求失败: {str(e)}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取用户信息请求处理失败: {str(e)}"
        )


@app.post("/lms/registerUser")
async def register_user(request: Request):
    """注册新用户，调用LMS的registerUser接口"""
    try:
        # 1. 从请求头获取authToken
        auth_token = request.headers.get('authToken')
        if not auth_token:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Unauthorized"
            )

        # 获取操作用户信息
        user_info = await get_user_info_from_token(auth_token)

        # 2. 获取请求体数据
        data = await request.json()
        new_user_name = data.get("userName", "未知用户")

        # 3. 调用LMS接口
        lms_register_url = f"{LMS_BASE_URL}/third/api/v1/userManagement/registerUser"
        headers = {
            "authToken": auth_token,
            "Content-Type": "application/json"
        }

        response = requests.post(lms_register_url, json=data, headers=headers)

        if response.status_code == 200:
            # 记录用户添加操作
            log_operation(
                operation_type="user_management",
                action="添加用户",
                user_id=user_info.get("userId"),
                user_name=user_info.get("userName"),
                target=new_user_name,
                status="success",
                details={
                    "new_user_data": data,
                    "response": response.json()
                }
            )

            return response.json()
        else:
            # 记录操作失败
            log_operation(
                operation_type="user_management",
                action="添加用户",
                user_id=user_info.get("userId"),
                user_name=user_info.get("userName"),
                target=new_user_name,
                status="failed",
                details={
                    "new_user_data": data,
                    "error": response.text[:200],
                    "status_code": response.status_code
                }
            )

            raise HTTPException(
                status_code=response.status_code,
                detail=f"LMS注册用户失败: {response.text}"
            )
    except Exception as e:
        logger.error(f"注册用户请求失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="注册用户请求处理失败"
        )


@app.post("/lms/deleteUser")
async def delete_user(request: Request):
    """删除用户，调用LMS的deleteUser接口"""
    try:
        # 1. 从请求头获取authToken
        auth_token = request.headers.get('authToken')
        if not auth_token:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Unauthorized"
            )

        # 2. 获取请求体数据
        data = await request.json()

        # 3. 调用LMS接口
        lms_delete_url = f"{LMS_BASE_URL}/third/api/v1/userManagement/deleteUser"
        headers = {
            "authToken": auth_token,
            "Content-Type": "application/json"
        }

        response = requests.post(lms_delete_url, json=data, headers=headers)

        if response.status_code == 200:
            return response.json()
        else:
            raise HTTPException(
                status_code=response.status_code,
                detail=f"LMS删除用户失败: {response.text}"
            )
    except Exception as e:
        logger.error(f"删除用户请求失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="删除用户请求处理失败"
        )


######################################### 历史数据 #########################################
# OUTPUT_ROOT = Path("/home/ubuntu/Projects/LeafDepot/output")
# 改用相对路径
OUTPUT_ROOT = Path(__file__).parent.parent.parent / "output"


def parse_task_date_from_filename(filename: str) -> Optional[datetime]:
    """从任务ID解析日期"""
    try:
        # 提取数字部分
        import re
        numbers = re.findall(r'\d+', filename)

        if numbers:
            # 取最长的数字串（可能是日期）
            longest_num = max(numbers, key=len)

            # 尝试解析为日期
            if len(longest_num) >= 8:
                date_str = longest_num[:8]
                year = int(date_str[:4])
                month = int(date_str[4:6])
                day = int(date_str[6:8])

                # 验证日期有效性
                if 2000 <= year <= 2100 and 1 <= month <= 12 and 1 <= day <= 31:
                    return datetime(year, month, day)

        # 如果无法解析，尝试从文件名中提取其他格式
        # 例如：HS2026011817 -> 2026-01-18
        pattern = r'(\d{4})(\d{2})(\d{2})'
        match = re.search(pattern, filename)
        if match:
            year, month, day = map(int, match.groups())
            return datetime(year, month, day)

        return None

    except Exception:
        return None


def is_task_expired(task_date: datetime) -> bool:
    """检查任务是否过期（超过6个月）"""
    now = datetime.now()
    six_months_ago = now.replace(year=now.year - (1 if now.month <= 6 else 0),
                                 month=(now.month - 6) % 12 or 12)

    if six_months_ago.month == 12 and now.month <= 6:
        six_months_ago = six_months_ago.replace(year=six_months_ago.year - 1)

    return task_date < six_months_ago


@app.get("/api/history/tasks")
async def get_history_tasks():
    """
    获取历史任务列表
    读取指定路径下的所有xlsx文件，解析为历史任务列表
    """
    try:
        print(f"文件位置: {OUTPUT_ROOT}")
        history_tasks_dir = OUTPUT_ROOT / "history_data"

        # 确保目录存在
        history_tasks_dir.mkdir(parents=True, exist_ok=True)

        # 查找所有xlsx文件
        xlsx_files = list(history_tasks_dir.glob("*.xlsx"))

        tasks = []

        for xlsx_file in xlsx_files:
            try:
                # 从文件名解析任务ID（去掉扩展名）
                task_id = xlsx_file.stem

                # 尝试从任务ID解析日期
                task_date = parse_task_date_from_filename(task_id)

                # 检查是否过期（超过6个月）
                is_expired = is_task_expired(task_date) if task_date else False

                tasks.append({
                    "taskId": task_id,
                    "taskDate": task_date.isoformat() if task_date else None,
                    "fileName": xlsx_file.name,
                    "isExpired": is_expired,
                    "filePath": str(xlsx_file.relative_to(OUTPUT_ROOT))
                })
            except Exception as e:
                logger.error(f"解析历史任务文件失败 {xlsx_file.name}: {str(e)}")
                continue

        # 按日期倒序排序（最新的在前）
        tasks.sort(key=lambda x: x.get("taskDate", ""), reverse=True)

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "code": 200,
                "message": "获取历史任务成功",
                "data": {
                    "tasks": tasks,
                    "total": len(tasks)
                }
            }
        )

    except Exception as e:
        logger.error(f"获取历史任务列表失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取历史任务列表失败: {str(e)}"
        )


@app.get("/api/history/task/{task_id}")
async def get_history_task_details(task_id: str):
    """
    获取历史任务详情
    读取与任务编号同名的xlsx文件，解析其中数据

    Args:
        task_id: 任务编号，如 HS2026011817
    """
    try:
        history_tasks_dir = OUTPUT_ROOT / "history_data"

        # 查找对应的xlsx文件
        xlsx_file = history_tasks_dir / f"{task_id}.xlsx"

        if not xlsx_file.exists():
            # 尝试查找其他可能的扩展名
            possible_files = list(history_tasks_dir.glob(f"{task_id}.*"))
            if possible_files:
                xlsx_file = possible_files[0]
            else:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"历史任务文件不存在: {task_id}.xlsx"
                )

        # 检查文件是否存在
        if not xlsx_file.exists():
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"文件不存在: {xlsx_file}"
            )

        logger.info(f"读取Excel文件: {xlsx_file}")

        # 方法1: 使用openpyxl读取Excel文件
        try:
            import openpyxl
            logger.info("使用openpyxl读取Excel文件")

            # 加载工作簿
            workbook = openpyxl.load_workbook(str(xlsx_file), data_only=True)

            # 获取第一个工作表
            sheet_name = workbook.sheetnames[0]
            worksheet = workbook[sheet_name]

            # 读取表头
            headers = []
            for cell in worksheet[1]:
                if cell.value:
                    headers.append(str(cell.value))

            logger.info(f"Excel表头: {headers}")

            # 读取数据行
            details = []
            row_index = 0

            for row in worksheet.iter_rows(min_row=2, values_only=True):
                row_index += 1

                # 跳过空行
                if all(cell is None for cell in row):
                    continue

                # 创建行字典
                row_dict = {}
                for i, header in enumerate(headers):
                    if i < len(row):
                        row_dict[header] = row[i]

                # 规范化数据格式
                detail = {
                    "序号": row_dict.get("序号") or row_index,
                    "品规名称": str(row_dict.get("品规名称", "")),
                    "储位名称": str(row_dict.get("储位名称", "")),
                    "实际品规": str(row_dict.get("实际品规", row_dict.get("品规名称", ""))),
                    "库存数量": int(row_dict.get("库存数量", 0) or 0),
                    "实际数量": int(row_dict.get("实际数量", 0) or 0),
                    "差异": str(row_dict.get("差异", "一致")),
                    "照片1路径": str(row_dict.get("照片1路径", "/3D_CAMERA/MAIN.JPEG")),
                    "照片2路径": str(row_dict.get("照片2路径", "/3D_CAMERA/DEPTH.JPEG")),
                    "照片3路径": str(row_dict.get("照片3路径", "/SCAN_CAMERA_1/1.JPEG")),
                    "照片4路径": str(row_dict.get("照片4路径", "/SCAN_CAMERA_2/1.JPEG")),
                }

                # 计算差异（如果未提供）
                if detail["差异"] == "一致" and detail["库存数量"] != detail["实际数量"]:
                    diff = detail["实际数量"] - detail["库存数量"]
                    if diff > 0:
                        detail["差异"] = f"多{diff}件"
                    else:
                        detail["差异"] = f"少{abs(diff)}件"

                # 如果差异为空，计算差异
                elif not detail["差异"] or detail["差异"].strip() == "":
                    if detail["库存数量"] == detail["实际数量"]:
                        detail["差异"] = "一致"
                    else:
                        diff = detail["实际数量"] - detail["库存数量"]
                        if diff > 0:
                            detail["差异"] = f"多{diff}件"
                        else:
                            detail["差异"] = f"少{abs(diff)}件"

                details.append(detail)

            logger.info(f"成功读取 {len(details)} 条记录")

        except ImportError:
            logger.warning("openpyxl未安装，尝试使用pandas")

            # 方法2: 使用pandas读取Excel文件
            try:
                import pandas as pd
                logger.info("使用pandas读取Excel文件")

                # 读取Excel文件
                df = pd.read_excel(str(xlsx_file))

                details = []
                for index, row in df.iterrows():
                    detail = {
                        "序号": row.get("序号") or (index + 1),
                        "品规名称": str(row.get("品规名称", "")),
                        "储位名称": str(row.get("储位名称", "")),
                        "实际品规": str(row.get("实际品规", row.get("品规名称", ""))),
                        "库存数量": int(row.get("库存数量", 0) or 0),
                        "实际数量": int(row.get("实际数量", 0) or 0),
                        "差异": str(row.get("差异", "一致")),
                        "照片1路径": str(row.get("照片1路径", "/3D_CAMERA/MAIN.JPEG")),
                        "照片2路径": str(row.get("照片2路径", "/3D_CAMERA/DEPTH.JPEG")),
                        "照片3路径": str(row.get("照片3路径", "/SCAN_CAMERA_1/1.JPEG")),
                        "照片4路径": str(row.get("照片4路径", "/SCAN_CAMERA_2/1.JPEG")),
                    }

                    # 计算差异（如果未提供）
                    if detail["差异"] == "一致" and detail["库存数量"] != detail["实际数量"]:
                        diff = detail["实际数量"] - detail["库存数量"]
                        if diff > 0:
                            detail["差异"] = f"多{diff}件"
                        else:
                            detail["差异"] = f"少{abs(diff)}件"

                    # 如果差异为空，计算差异
                    elif not detail["差异"] or pd.isna(detail["差异"]):
                        if detail["库存数量"] == detail["实际数量"]:
                            detail["差异"] = "一致"
                        else:
                            diff = detail["实际数量"] - detail["库存数量"]
                            if diff > 0:
                                detail["差异"] = f"多{diff}件"
                            else:
                                detail["差异"] = f"少{abs(diff)}件"

                    details.append(detail)

                logger.info(f"成功读取 {len(details)} 条记录")

            except ImportError:
                logger.error("未找到Excel解析库，请安装openpyxl或pandas")
                raise HTTPException(
                    status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    detail="Excel解析库未安装，请安装openpyxl或pandas"
                )

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "code": 200,
                "message": "获取历史任务详情成功",
                "data": {
                    "taskId": task_id,
                    "fileName": xlsx_file.name,
                    "details": details,
                    "totalBins": len(details)
                }
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取历史任务详情失败 {task_id}: {str(e)}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取历史任务详情失败: {str(e)}"
        )


@app.get("/api/history/task/{task_id}/bin/{bin_location}")
async def get_history_bin_detail(task_id: str, bin_location: str):
    """
    获取历史任务中特定储位的详情

    Args:
        task_id: 任务编号
        bin_location: 储位名称
    """
    try:
        # 先获取整个任务的详情
        from fastapi import BackgroundTasks
        import asyncio

        # 创建模拟请求来调用上面的接口
        task_details_response = await get_history_task_details(task_id)

        if task_details_response.status_code != 200:
            raise HTTPException(
                status_code=task_details_response.status_code,
                detail=task_details_response.body.decode(
                ) if task_details_response.body else "获取任务详情失败"
            )

        # 解析响应数据
        response_data = json.loads(task_details_response.body.decode())
        if response_data["code"] != 200:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=response_data["message"]
            )

        # 查找特定储位
        details = response_data["data"]["details"]
        bin_detail = None
        for detail in details:
            if detail["储位名称"] == bin_location:
                bin_detail = detail
                break

        if not bin_detail:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"储位 {bin_location} 不存在于任务 {task_id} 中"
            )

        # 获取该储位的图片
        image_urls = []
        photo_fields = ["照片1路径", "照片2路径", "照片3路径", "照片4路径"]

        for i, photo_field in enumerate(photo_fields, 1):
            photo_path = bin_detail.get(photo_field, "")
            if photo_path:
                # 构建完整的图片路径
                image_url = f"/api/history/image?taskNo={task_id}&binLocation={bin_location}&cameraType={photo_path.split('/')[1]}&filename={Path(photo_path).name}"
                image_urls.append({
                    "index": i,
                    "path": photo_path,
                    "url": image_url,
                    "cameraType": photo_path.split('/')[1] if '/' in photo_path else "未知",
                    "filename": Path(photo_path).name
                })

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "code": 200,
                "message": "获取储位详情成功",
                "data": {
                    "taskId": task_id,
                    "binLocation": bin_location,
                    "detail": bin_detail,
                    "imageUrls": image_urls,
                    "hasImages": len(image_urls) > 0
                }
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取历史任务储位详情失败 {task_id}/{bin_location}: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取储位详情失败: {str(e)}"
        )


@app.get("/api/history/image")
async def get_history_image(
    taskNo: str,
    binLocation: str,
    cameraType: str,
    filename: str,
    source: str = "output"  # 新增参数：source可以是"output"、"capture_img"或"history"
):
    """
    获取盘点任务中的图片

    Args:
        taskNo: 任务编号
        binLocation: 储位名称
        cameraType: 相机类型
        filename: 文件名
    """
    try:
        project_root = OUTPUT_ROOT

        # 构建基本路径
        base_path = project_root / taskNo / binLocation / cameraType

        # 定义可能的图片扩展名
        image_extensions = ['.jpg', '.jpeg', '.png', '.bmp', '.gif', '.webp']

        image_path = None

        # 尝试查找不同扩展名的文件
        for ext in image_extensions:
            test_path = base_path / f"{filename}{ext}"
            if test_path.exists():
                image_path = test_path
                break

        # 如果还没找到，尝试查找大写扩展名的文件
        if not image_path:
            for ext in image_extensions:
                test_path = base_path / f"{filename}{ext.upper()}"
                if test_path.exists():
                    image_path = test_path
                    break

        if not image_path:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"图片不存在: {filename} (路径: {base_path})"
            )

        # 读取图片文件
        with open(image_path, "rb") as f:
            image_data = f.read()

        # 根据文件扩展名确定媒体类型
        image_ext = image_path.suffix.lower()
        if image_ext in ['.jpg', '.jpeg']:
            media_type = "image/jpeg"
        elif image_ext == '.png':
            media_type = "image/png"
        elif image_ext == '.bmp':
            media_type = "image/bmp"
        elif image_ext == '.gif':
            media_type = "image/gif"
        elif image_ext == '.webp':
            media_type = "image/webp"
        else:
            # 默认使用jpeg
            media_type = "image/jpeg"
            logger.warning(f"未知的图片扩展名 {image_ext}，使用默认媒体类型")

        # 记录成功信息
        logger.info(
            f"成功返回图片: {image_path.name}, 大小: {len(image_data)} bytes, 类型: {media_type}")

        # 添加 CORS 头
        headers = {
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Methods": "GET, OPTIONS",
            "Access-Control-Allow-Headers": "*",
            "Access-Control-Expose-Headers": "*",
            "Cache-Control": "public, max-age=3600"
        }

        return Response(content=image_data, media_type=media_type, headers=headers)

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取图片失败: {str(e)}")
        # 返回 404 时也要添加 CORS 头
        headers = {
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Methods": "GET, OPTIONS",
            "Access-Control-Allow-Headers": "*"
        }
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"图片不存在: {filename}",
            headers=headers
        )


# 操作记录接口
@app.get("/api/operation/logs/recent")
async def get_recent_operation_logs(limit: int = 5):
    """
    获取最近的操作记录（默认最近5条，6个月内）

    Args:
        limit: 返回记录数量，默认5条
    """
    try:
        operations = get_recent_operations(limit=limit, days=180)

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "code": 200,
                "message": "获取操作记录成功",
                "data": {
                    "operations": operations,
                    "total": len(operations),
                    "limit": limit
                }
            }
        )

    except Exception as e:
        logger.error(f"获取操作记录失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取操作记录失败: {str(e)}"
        )


@app.get("/api/operation/logs/all")
async def get_all_operation_logs(
    days: int = 180,
    operation_type: Optional[str] = None,
    status: Optional[str] = None
):
    """
    获取所有操作记录（可筛选）

    Args:
        days: 查询天数，默认180天（6个月）
        operation_type: 操作类型筛选
        status: 状态筛选
    """
    try:
        operations = get_all_operations(days=days)

        # 应用筛选
        filtered_operations = operations

        if operation_type:
            filtered_operations = [op for op in filtered_operations
                                   if op.get("operation_type") == operation_type]

        if status:
            filtered_operations = [op for op in filtered_operations
                                   if op.get("status") == status]

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "code": 200,
                "message": "获取操作记录成功",
                "data": {
                    "operations": filtered_operations,
                    "total": len(filtered_operations),
                    "filtered_total": len(operations),
                    "days": days
                }
            }
        )

    except Exception as e:
        logger.error(f"获取操作记录失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取操作记录失败: {str(e)}"
        )


# 5. 历史数据清理接口
class HistoryCleanupRequest(BaseModel):
    """历史数据清理请求模型"""
    cutoff_date: Optional[str] = None  # 截止日期，格式: YYYY-MM-DD
    days: Optional[int] = 180  # 保留天数，默认180天（6个月）


@app.post("/api/history/cleanup")
async def cleanup_history_data(request: Request):
    """
    清理历史数据
    删除指定天数之前的历史任务文件

    Request body:
        cutoff_date: 截止日期（可选），格式: YYYY-MM-DD
        days: 保留天数（可选），默认180天（6个月）

    Returns:
        code: 状态码
        message: 消息
        data: 清理结果
            cleaned_count: 清理的文件数量
            cutoff_date: 实际使用的截止日期
            retention_days: 实际保留天数
            cleaned_files: 清理的文件列表
    """
    try:
        # 解析请求参数
        data = await request.json()

        # 获取参数
        cutoff_date_str = data.get("cutoff_date")
        days = data.get("days", 180)

        # 计算截止日期
        if cutoff_date_str:
            # 使用提供的截止日期
            try:
                cutoff_date = datetime.strptime(cutoff_date_str, "%Y-%m-%d")
            except ValueError:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"截止日期格式错误，应为 YYYY-MM-DD，实际为: {cutoff_date_str}"
                )
        else:
            # 使用天数计算截止日期
            cutoff_date = datetime.now() - timedelta(days=days)

        cutoff_date_str = cutoff_date.strftime("%Y-%m-%d")
        logger.info(f"开始清理历史数据，截止日期: {cutoff_date_str}")

        # 获取历史数据目录
        history_tasks_dir = OUTPUT_ROOT / "history_data"

        if not history_tasks_dir.exists():
            logger.warning(f"历史数据目录不存在: {history_tasks_dir}")
            return JSONResponse(
                status_code=status.HTTP_200_OK,
                content={
                    "code": 200,
                    "message": "历史数据目录不存在，无需清理",
                    "data": {
                        "cleaned_count": 0,
                        "cutoff_date": cutoff_date_str,
                        "retention_days": days,
                        "cleaned_files": []
                    }
                }
            )

        # 查找所有历史数据文件
        all_files = list(history_tasks_dir.glob("*.xlsx"))
        all_files.extend(history_tasks_dir.glob("*.xls"))

        cleaned_files = []
        cleaned_count = 0

        for file_path in all_files:
            try:
                # 从文件名解析任务日期
                task_date = parse_task_date_from_filename(file_path.stem)

                if task_date is None:
                    # 如果无法解析日期，跳过该文件
                    logger.warning(f"无法从文件名解析日期: {file_path.name}")
                    continue

                # 检查是否需要删除
                if task_date < cutoff_date:
                    # 删除文件
                    file_path.unlink()
                    cleaned_files.append({
                        "filename": file_path.name,
                        "task_date": task_date.isoformat(),
                        "file_path": str(file_path)
                    })
                    cleaned_count += 1
                    logger.info(
                        f"已删除历史数据文件: {file_path.name} (日期: {task_date})")

            except Exception as e:
                logger.error(f"处理文件失败 {file_path.name}: {str(e)}")
                continue

        # 记录清理操作
        client_host = request.client.host if request.client else "unknown"
        log_operation(
            operation_type="system_cleanup",
            action="清理历史数据",
            status="success",
            ip_address=client_host,
            details={
                "cleaned_count": cleaned_count,
                "cutoff_date": cutoff_date_str,
                "retention_days": days,
                "cleaned_files": cleaned_files,
                "request_ip": client_host
            }
        )

        logger.info(f"历史数据清理完成，共删除 {cleaned_count} 个文件")

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "code": 200,
                "message": f"历史数据清理完成，共删除 {cleaned_count} 个文件",
                "data": {
                    "cleaned_count": cleaned_count,
                    "cutoff_date": cutoff_date_str,
                    "retention_days": days,
                    "cleaned_files": cleaned_files
                }
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"清理历史数据失败: {str(e)}")

        # 记录失败操作
        client_host = request.client.host if request.client else "unknown"
        log_operation(
            operation_type="system_cleanup",
            action="清理历史数据",
            status="failed",
            ip_address=client_host,
            details={
                "error": str(e),
                "request_ip": client_host
            }
        )

        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"清理历史数据失败: {str(e)}"
        )


# 6. 添加历史数据清理记录函数
async def log_history_cleanup(cleaned_count: int, cutoff_date: str, details: Optional[Dict] = None):
    """记录历史数据清理操作"""
    log_operation(
        operation_type="system_cleanup",
        action="清理历史数据",
        status="success",
        details={
            "cleaned_count": cleaned_count,
            "cutoff_date": cutoff_date,
            "retention_days": 180,
            ** (details or {})
        }
    )


if __name__ == "__main__":
    # 确保日志配置正确（已经在文件开头配置，这里不再重复配置）
    logger.info("🚀 Gateway服务启动")
    logger.info("📡 API地址: http://0.0.0.0:8000")
    logger.info("📚 API文档: http://localhost:8000/docs")
    logger.info("🔍 测试页面: http://localhost:8080/test_detect.html")
    logger.info(f"📝 日志文件: {_log_filename}")
    # 修复后的代码
    logger.info(
        f"📝 前端日志文件: {_debug_log_dir}/frontend_{datetime.now().strftime('%Y%m%d')}.log")
    logger.info("\n按 Ctrl+C 停止服务\n")

    uvicorn.run(app, host="0.0.0.0", port=8000, log_level="info")
